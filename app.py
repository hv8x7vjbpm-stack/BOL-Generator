import os, re, json, tempfile
from datetime import datetime
from flask import Flask, render_template, request, send_file, jsonify
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Frame
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfWriter
import pdfplumber

app = Flask(__name__)
BOL_STORE_PATH = os.path.join(os.path.dirname(__file__), "bol_store.json")


# ── Page constants ────────────────────────────────────────────────────────────
PW, PH = letter          # 612, 792 pt
ML = MR = MT = MB = 36  # 0.5 inch margins
TW = PW - ML - MR       # 540pt usable width
TX = ML                  # table left edge x
TY_TOP = PH - MT         # table top y (ReportLab y=0 is bottom)

# ── Colours ───────────────────────────────────────────────────────────────────
BLACK  = colors.black
WHITE  = colors.white
GREY   = colors.HexColor("#E6E6E6")

# ── Column widths (pt, scaled from exact EMU measurements) ───────────────────
SCALE  = TW / 549.2      # 549.2pt = total docx table width

C_L    = 273.6 * SCALE   # Left half  (Ship From side)
C_R    = 275.6 * SCALE   # Right half (BOL Number side)
C_PRO  = C_R / 2         # PRO NUMBER
C_TRL  = C_R / 2         # TRAILER

PO_W = [e/12700*SCALE for e in [2674620, 800100, 468630, 731520, 2299970]]
COM_W = [e/12700*SCALE for e in [285750,586740,349885,492125,457200,428625,3181985,596265,596265]]

# Value declaration / COD split
SCALE2 = TW / ((4.1125 + 3.5153) * inch / SCALE)  # proportional
C_VAL  = TW * 4.1125 / (4.1125 + 3.5153)
C_COD  = TW * 3.5153 / (4.1125 + 3.5153)

# Received / sig split
C_REC  = TW * 3.3625 / (3.3625 + 4.2653)
C_SIG_R = TW * 4.2653 / (3.3625 + 4.2653)

# Bottom sig 4-col
SIG4 = [TW * x / (2.3750+0.9875+1.9139+2.3514) for x in [2.3750,0.9875,1.9139,2.3514]]

# ── Row heights (pt, from real BOL EMU measurements) ─────────────────────────
RH = [
    23.05,   # 0  title bar (fixed)
    14.40,   # 1  Ship From label (fixed)
    115.40,  # 2  Ship From content
    14.40,   # 3  Ship To label (fixed)
    56.10,   # 4  Ship To / Carrier content
    14.40,   # 5  Third Party label (fixed)
    56.10,   # 6  Third Party content / PRO / Trailer
    31.17,   # 7  TMS ID / Freight terms
    14.40,   # 8  DO NOT STACK / Master bill (fixed)
    14.40,   # 9  Customer Order header (fixed)
    14.40,   # 10 PO table header (fixed)
    14.03,   # 11-18: 8 PO data rows
    14.03, 14.03, 14.03, 14.03, 14.03, 14.03, 14.03,
    43.64,   # 19 Commodity header
    22.44,   # 20 Commodity data
    22.44,   # 21 Value declaration / COD
    10.80,   # 22 Liability note (fixed)
    56.10,   # 23 Received text / carrier
    84.15,   # 24 Signature row
]

# ── Paragraph styles ──────────────────────────────────────────────────────────
PAD = 2  # cell padding pts

def sty(size=7, bold=False, align=TA_LEFT, color=BLACK, leading=None):
    fn = "Helvetica-Bold" if bold else "Helvetica"
    return ParagraphStyle("x", fontName=fn, fontSize=size,
                          leading=leading or size*1.25,
                          textColor=color, alignment=align,
                          wordWrap='CJK', spaceBefore=0, spaceAfter=0)

S_LABEL  = sty(6.5)
S_LABELC = sty(6.5, align=TA_CENTER)
S_NORM   = sty(7.5)
S_NORMC  = sty(7.5, align=TA_CENTER)
S_BOLD   = sty(7.5, bold=True)
S_BOLDC  = sty(7.5, bold=True, align=TA_CENTER)
S_BIG    = sty(10)
S_BIGB   = sty(10, bold=True)
S_TITLE  = sty(8, bold=True, align=TA_CENTER, color=WHITE)
S_TINY   = sty(5.5)
S_MICRO  = sty(4.5)
S_TINYC  = sty(5.5, align=TA_CENTER)
S_SMALL  = sty(6)
S_SMALLB = sty(6, bold=True)
S_DONT   = sty(7, bold=True)

def para(text, style):
    safe = str(text).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    return Paragraph(safe, style)

# ── Drawing helpers ───────────────────────────────────────────────────────────

def draw_rect(c, x, y, w, h, fill=None, stroke=True):
    """Draw rectangle. y is TOP of cell in page coordinates."""
    if fill:
        c.setFillColor(fill)
        c.rect(x, y - h, w, h, fill=1, stroke=0)
    if stroke:
        c.setStrokeColor(BLACK)
        c.setLineWidth(0.5)
        c.rect(x, y - h, w, h, fill=0, stroke=1)


def draw_cell(c, x, y, w, h, paragraphs, fill=None, pad=PAD):
    """Draw a table cell with optional fill and text content."""
    draw_rect(c, x, y, w, h, fill=fill)
    if not paragraphs:
        return
    if not isinstance(paragraphs, list):
        paragraphs = [paragraphs]
    # Use a Frame to flow paragraphs
    frame = Frame(
        x + pad, y - h + pad,
        w - pad*2, h - pad*2,
        leftPadding=0, rightPadding=0,
        topPadding=0, bottomPadding=0,
        showBoundary=0
    )
    story = []
    for p in paragraphs:
        if p is not None:
            story.append(p)
    frame.addFromList(story, c)


def draw_vline(c, x, y_top, y_bot):
    c.setStrokeColor(BLACK)
    c.setLineWidth(0.5)
    c.line(x, y_top, x, y_bot)


# ── Main generator ────────────────────────────────────────────────────────────

def generate_bol_pdf(data, out_path):
    bol_number     = data.get("bol_number", "")
    address        = data.get("address", "")
    carrier        = data.get("carrier", "")
    pro_number     = data.get("pro_number", "")
    shipment       = data.get("shipment", "")
    po_numbers     = data.get("po_numbers", [])
    pallets_per_po = data.get("pallets_per_po", [])
    total_pallets  = data.get("total_pallets", "")
    total_weight   = data.get("total_weight", "")

    c = canvas.Canvas(out_path, pagesize=letter)

    # Precompute row y positions (y = top of each row)
    y = []
    cur = TY_TOP
    for h in RH:
        y.append(cur)
        cur -= h

    # ── R0: Title bar ─────────────────────────────────────────────────────────
    r, h = 0, RH[0]
    # Three cells
    w0 = TW * 1.3368 / 7.6278
    w1 = TW * 4.5257 / 7.6278
    w2 = TW * 1.7653 / 7.6278
    draw_cell(c, TX,        y[r], w0, h, para("", S_TITLE),                                      fill=BLACK)
    draw_cell(c, TX+w0,     y[r], w1, h, para("Bill of Lading \u2013 Short Form \u2013 Non-Negotiable", S_TITLE), fill=BLACK)
    draw_cell(c, TX+w0+w1,  y[r], w2, h, para("Page 1 of 1", S_TITLE),                          fill=BLACK)

    # ── R1: Ship From label | BOL Number label ────────────────────────────────
    r, h = 1, RH[1]
    draw_cell(c, TX,      y[r], C_L, h, para("Ship From", S_LABEL), fill=GREY)
    draw_cell(c, TX+C_L,  y[r], C_R, h, para("Bill of Lading Number:", S_LABEL))

    # ── R2: Ship From content | BOL Number value ──────────────────────────────
    r, h = 2, RH[2]
    draw_cell(c, TX, y[r], C_L, h, [
        para("JACKSON POTTERY INC", S_NORM),
        para("2146 EMPIRE CENTRAL", S_NORM),
        para("DALLAS, TX 75235", S_NORM),
        para("214-974-0679", S_NORM),
    ])
    draw_cell(c, TX+C_L, y[r], C_R, h, para(bol_number, S_BIG))

    # ── R3: Ship To label | Carrier Name label ────────────────────────────────
    r, h = 3, RH[3]
    draw_cell(c, TX,      y[r], C_L, h, para("Ship To", S_LABEL), fill=GREY)
    draw_cell(c, TX+C_L,  y[r], C_R, h, para("Carrier Name:", S_LABEL))

    # ── R4: Ship To content | Carrier value ───────────────────────────────────
    r, h = 4, RH[4]
    addr_lines = address.replace("\r", "").split("\n")
    ship_to = [para("HOME DEPOT \u2013 Store", S_BIGB)]
    for line in addr_lines:
        if line.strip():
            ship_to.append(para(line.strip(), S_BOLD))
    draw_cell(c, TX,      y[r], C_L, h, ship_to)
    draw_cell(c, TX+C_L,  y[r], C_R, h, para(carrier, S_BIGB))

    # ── R5: Third Party label | PRO label | Trailer label ────────────────────
    r, h = 5, RH[5]
    draw_cell(c, TX,             y[r], C_L,   h, para("Third Party Freight Charges Bill to", S_LABEL), fill=GREY)
    draw_cell(c, TX+C_L,         y[r], C_PRO, h, para("PRO NUMBER", S_LABEL), fill=GREY)
    draw_cell(c, TX+C_L+C_PRO,   y[r], C_TRL, h, para("TRAILER / SEAL NUMBER", S_LABEL), fill=GREY)

    # ── R6: Third Party content | PRO value | Trailer value ──────────────────
    r, h = 6, RH[6]
    draw_cell(c, TX, y[r], C_L, h, [
        para("HOMEDEPOT.COM/ATTN: FREIGHT PAYABLES", S_NORM),
        para("2455 PACES FERRY RD", S_NORM),
        para("ATLANTA, GA 30339", S_NORM),
    ])
    draw_cell(c, TX+C_L,       y[r], C_PRO, h, para(pro_number, S_NORM))
    draw_cell(c, TX+C_L+C_PRO, y[r], C_TRL, h, para("", S_NORM))

    # ── R7: TMS ID left | Freight Charge Terms right ─────────────────────────
    r, h = 7, RH[7]
    draw_cell(c, TX, y[r], C_L, h, [
        para("TMS ID NUMBER", S_LABEL),
        para(shipment, S_BIG),
    ])
    draw_cell(c, TX+C_L, y[r], C_R, h, [
        para("Freight Charge Terms (Freight charges are prepaid unless marked otherwise):", S_SMALL),
        para("Prepaid \u2610  Collect \u2610  3rd Party  x", S_SMALL),
    ])

    # ── R8: DO NOT STACK left | Master bill right ────────────────────────────
    r, h = 8, RH[8]
    draw_cell(c, TX,     y[r], C_L, h, para("DO NOT STACK PALLETS", S_DONT))
    draw_cell(c, TX+C_L, y[r], C_R, h, para("\u2610 Master bill of lading with attached underlying bills of lading.", S_SMALL))

    # ── R9: Customer Order Information ───────────────────────────────────────
    r, h = 9, RH[9]
    draw_cell(c, TX, y[r], TW, h, para("Customer Order Information", S_TITLE), fill=BLACK)

    # ── R10: PO header ────────────────────────────────────────────────────────
    r, h = 10, RH[10]
    px = TX
    hdrs = [
        (para("SPECIAL INSTRUCTIONS\nPO NUMBERS", S_LABEL), PO_W[0]),
        (para("# of Pallets", S_LABELC), PO_W[1]),
        (para("", S_LABEL), PO_W[2]),
        (para("Pallet/Slip\n(circle one)", S_LABELC), PO_W[3]),
        (para("Additional Shipper Information", S_LABEL), PO_W[4]),
    ]
    for p_text, pw in hdrs:
        draw_cell(c, px, y[r], pw, h, p_text, fill=GREY)
        px += pw

    # ── R11-R18: PO data rows ─────────────────────────────────────────────────
    for i in range(8):
        r = 11 + i
        h = RH[r]
        pov = po_numbers[i]     if i < len(po_numbers)     else ""
        pav = pallets_per_po[i] if i < len(pallets_per_po) else ""
        px = TX
        for content, pw in [
            (para(pov, S_NORM), PO_W[0]),
            (para(pav, S_NORMC), PO_W[1]),
            (para("", S_NORM), PO_W[2]),
            (para("", S_NORM), PO_W[3]),
            (para("", S_NORM), PO_W[4]),
        ]:
            draw_cell(c, px, y[r], pw, h, content)
            px += pw

    # ── R19: Commodity header ─────────────────────────────────────────────────
    r, h = 19, RH[19]
    px = TX
    com_hdrs = [
        (para("Qty", S_LABELC), COM_W[0]),
        (para("Type", S_LABEL), COM_W[1]),
        (para("Qty", S_LABELC), COM_W[2]),
        (para("Type", S_LABEL), COM_W[3]),
        (para("Weight", S_LABEL), COM_W[4]),
        (para("HM(X)", S_LABELC), COM_W[5]),
        (para("Commodity Description  Commodities requiring special or additional "
              "care or attention in handling or stowing must be so marked and packaged "
              "as to ensure safe transportation with ordinary care. "
              "See Section 2(e) of NMFC item 360", S_TINY), COM_W[6]),
        (para("NMFC No.", S_LABELC), COM_W[7]),
        (para("Class", S_LABELC), COM_W[8]),
    ]
    for p_text, pw in com_hdrs:
        draw_cell(c, px, y[r], pw, h, p_text, fill=GREY)
        px += pw

    # ── R20: Commodity data ───────────────────────────────────────────────────
    r, h = 20, RH[20]
    px = TX
    com_data = [
        (para(total_pallets, S_NORMC), COM_W[0]),
        (para("PALLETS", S_NORM), COM_W[1]),
        (para("", S_NORM), COM_W[2]),
        (para("", S_NORM), COM_W[3]),
        (para(total_weight, S_NORMC), COM_W[4]),
        (para("", S_NORM), COM_W[5]),
        (para("CERAMIC, CHINA, EARTHENWARE, PORCELAIN OR STONEWARE/ POTTERY", S_NORM), COM_W[6]),
        (para("47500-12", S_NORMC), COM_W[7]),
        (para("55", S_NORMC), COM_W[8]),
    ]
    for p_text, pw in com_data:
        draw_cell(c, px, y[r], pw, h, p_text)
        px += pw

    # ── R21: Value declaration | COD ─────────────────────────────────────────
    r, h = 21, RH[21]
    draw_cell(c, TX, y[r], C_VAL, h,
        para('Where the rate is dependent on value, shippers are required to state '
             'specifically in writing the agreed or declared value of the property as '
             'follows: "The agreed or declared value of the property is specifically '
             'stated by the shipper to be not exceeding _______________ per _______________.', S_MICRO))
    draw_cell(c, TX+C_VAL, y[r], C_COD, h, [
        para("COD Amount: $", S_MICRO),
        para("Fee terms: Collect \u2610  Prepaid \u2610  Customer check acceptable \u2610", S_MICRO),
    ])

    # ── R22: Liability note ───────────────────────────────────────────────────
    r, h = 22, RH[22]
    draw_cell(c, TX, y[r], TW, h,
        para("Note: Liability limitation for loss or damage in this shipment may be "
             "applicable. See 49 USC \u00a7 14706(c)(1)(A) and (B).", S_TINY))

    # ── R23: Received text | Carrier payment ─────────────────────────────────
    r, h = 23, RH[23]
    draw_cell(c, TX, y[r], C_REC, h,
        para("Received, subject to individually determined rates or contracts that "
             "have been agreed upon in writing between the carrier and shipper, if "
             "applicable, otherwise to the rates, classifications, and rules that have "
             "been established by the carrier and are available to the shipper, on "
             "request, and to all applicable state and federal regulations.", S_TINY))
    draw_cell(c, TX+C_REC, y[r], C_SIG_R, h, [
        para("The carrier shall not make delivery of this shipment without payment "
             "of charges and all other lawful fees.", S_TINY),
        para("Shipper Signature: _________________________", S_TINY),
    ])

    # ── R24: Signature 4-col row ──────────────────────────────────────────────
    r, h = 24, RH[24]
    px = TX
    sig_contents = [
        [para("Shipper Signature/Date", S_LABEL),
         para(" ", S_TINY),
         para("This is to certify that the above-named materials are properly "
              "classified, packaged, marked, and labeled, and are in proper condition "
              "for transportation according to the applicable regulations of the DOT.", S_TINY)],
        [para("Trailer Loaded:", S_LABEL),
         para("x By shipper", S_SMALL),
         para("\u2610 By driver", S_SMALL),
         para("Trailer Counted", S_SMALLB),
         para("x By shipper", S_SMALL),
         para("\u2610 By driver", S_SMALL)],
        [para("Freight Counted:", S_LABEL),
         para("x By shipper", S_SMALL),
         para("\u2610 By driver/pallets said to contain", S_SMALL),
         para("\u2610 By driver/pieces", S_SMALL)],
        [para("Carrier Signature/Pickup Date", S_LABEL),
         para(" ", S_TINY),
         para("Carrier acknowledges receipt of packages and required placards. "
              "Carrier certifies emergency response information was made available "
              "and/or carrier has the DOT emergency response guidebook or equivalent "
              "documentation in the vehicle. Property described above is received in "
              "good order, except as noted.", S_TINY)],
    ]
    for content, sw in zip(sig_contents, SIG4):
        draw_cell(c, px, y[r], sw, h, content)
        px += sw

    c.save()


# ── Store helpers ─────────────────────────────────────────────────────────────
def load_store():
    if os.path.exists(BOL_STORE_PATH):
        with open(BOL_STORE_PATH,"r") as f: return json.load(f)
    return []

def save_store(bols):
    with open(BOL_STORE_PATH,"w") as f: json.dump(bols,f,indent=2)

def get_next_id(bols):
    return 1 if not bols else max(b["id"] for b in bols)+1

def clean_bol_data(d):
    out={}
    for k,v in d.items():
        if isinstance(v,str):   out[k]="" if v.strip().upper()=="N/A" else v
        elif isinstance(v,list):out[k]=["" if x.strip().upper()=="N/A" else x for x in v]
        else: out[k]=v
    return out


def generate_bols_pdf(bols_data):
    if len(bols_data) == 1:
        tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        tmp.close()
        generate_bol_pdf(bols_data[0], tmp.name)
        return tmp.name
    writer = PdfWriter()
    tmps = []
    try:
        for bol in bols_data:
            t = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            t.close(); tmps.append(t.name)
            generate_bol_pdf(bol, t.name)
            writer.append(t.name)
        out = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
        out.close()
        with open(out.name, "wb") as f: writer.write(f)
        return out.name
    finally:
        for p2 in tmps:
            try: os.unlink(p2)
            except: pass


def extract_bol_from_pdf(pdf_path):
    result = {"bol_number":"","address":"","carrier":"","pro_number":"",
              "shipment":"","po_numbers":[],"pallets_per_po":[],"total_pallets":"","total_weight":""}
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]; full_text = page.extract_text() or ""; tables = page.extract_tables()
    lines = full_text.split("\n")
    for i,line in enumerate(lines):
        if "bill of lading number" in line.lower():
            parts = line.split(":")
            v = parts[-1].strip() if len(parts)>1 else ""
            result["bol_number"] = v if v not in ("","INPUT") else (lines[i+1].strip() if i+1<len(lines) else "")
            break
    addr_lines, in_addr = [], False
    for line in lines:
        if "ship to" in line.lower() and not in_addr: in_addr=True; continue
        if in_addr:
            if any(x in line.lower() for x in ("third party","carrier name","pro number","freight")): break
            s = line.strip()
            if s and s not in ("INPUT","HOME DEPOT"): addr_lines.append(s)
    result["address"] = "\n".join(addr_lines[:4])
    for i,line in enumerate(lines):
        if "carrier name" in line.lower():
            parts = line.split(":")
            v = parts[-1].strip() if len(parts)>1 else ""
            if v and v not in ("","INPUT"): result["carrier"]=v
            else:
                for j in range(1,4):
                    if i+j<len(lines) and lines[i+j].strip() not in ("","INPUT"):
                        result["carrier"]=lines[i+j].strip(); break
            break
    for i,line in enumerate(lines):
        if "pro number" in line.lower():
            parts = re.split(r"pro number[:\s]*",line,flags=re.IGNORECASE)
            v = parts[-1].strip() if len(parts)>1 else ""
            result["pro_number"] = v if v not in ("","INPUT") else (lines[i+1].strip() if i+1<len(lines) else "")
            break
    for i,line in enumerate(lines):
        if "tms id" in line.lower():
            m = re.search(r"tms id number[:\s]+([\w\-]+)",line,re.IGNORECASE)
            if m: result["shipment"]=m.group(1)
            else:
                for j in range(1,4):
                    if i+j<len(lines):
                        v=lines[i+j].strip()
                        if v and v not in ("INPUT","DO NOT STACK PALLETS",""): result["shipment"]=v; break
            break
    po_numbers, pallets_per_po = [], []
    for table in tables:
        for row in (table or []):
            if not row: continue
            rt = [str(c).strip() if c else "" for c in row]
            for ci,cell in enumerate(rt):
                if re.match(r"^(PO[-\s]?\d+|\d{5,})$",cell,re.IGNORECASE):
                    po_numbers.append(cell)
                    pal = next((c2 for c2 in rt[ci+1:] if c2 and re.match(r"^\d+$",c2)),"")
                    pallets_per_po.append(pal)
    seen,upos,upals = set(),[],[]
    for i,po in enumerate(po_numbers):
        if po not in seen:
            seen.add(po); upos.append(po)
            upals.append(pallets_per_po[i] if i<len(pallets_per_po) else "")
    result["po_numbers"]=upos; result["pallets_per_po"]=upals
    half = lines[len(lines)//2:]
    for line in half:
        if "pallet" in line.lower() and not result["total_pallets"]:
            m = re.search(r"\b(\d+)\b",line)
            if m: result["total_pallets"]=m.group(1)
    for line in half:
        if not result["total_weight"]:
            m = re.search(r"(\d{3,6})\s*(?:lbs?)?",line,re.IGNORECASE)
            if m and int(m.group(1))>100: result["total_weight"]=m.group(1)
    result["po_numbers"]=[p2 for p2 in result["po_numbers"] if p2.strip()]
    result["pallets_per_po"]=result["pallets_per_po"][:len(result["po_numbers"])]
    return result


def build_excel_shortage(bols):
    wb=openpyxl.Workbook(); ws_all=wb.active; ws_all.title="All Orders"
    hf=PatternFill("solid",fgColor="1A365D"); hfont=Font(bold=True,color="FFFFFF",name="Arial",size=10)
    af=PatternFill("solid",fgColor="EBF2FA"); bs=Side(style="thin",color="CCCCCC")
    cb=Border(left=bs,right=bs,top=bs,bottom=bs)
    ctr=Alignment(horizontal="center",vertical="center",wrap_text=True)
    lft=Alignment(horizontal="left",  vertical="center",wrap_text=True)
    def make_hdr(ws,hdrs,widths):
        ws.append(hdrs)
        for col in range(1,len(hdrs)+1):
            c=ws.cell(1,col); c.fill=hf; c.font=hfont; c.alignment=ctr; c.border=cb
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
        ws.row_dimensions[1].height=28; ws.freeze_panes="A2"
    def style_cells(ws,ri,ncols,fill=None,bold=False):
        for col in range(1,ncols+1):
            c=ws.cell(ri,col)
            c.fill=fill if fill else (af if ri%2==0 else PatternFill())
            c.font=Font(name="Arial",size=10,bold=bold); c.alignment=ctr; c.border=cb
    make_hdr(ws_all,["Order #","BOL Number","Ship To Address","Carrier","PRO Number",
        "TMS / Shipment","PO Numbers","Pallets per PO","Total Pallets","Total Weight (lbs)","Date Added"],
        [8,14,30,18,14,16,20,14,12,16,18])
    for ri,bol in enumerate(bols,2):
        ws_all.append([bol.get("id",""),bol.get("bol_number",""),bol.get("address",""),
            bol.get("carrier",""),bol.get("pro_number",""),bol.get("shipment",""),
            "\n".join(bol.get("po_numbers",[])),"\n".join(bol.get("pallets_per_po",[])),
            bol.get("total_pallets",""),bol.get("total_weight",""),bol.get("date_added","")])
        style_cells(ws_all,ri,11); ws_all.cell(ri,3).alignment=lft
    ws_po=wb.create_sheet("By PO Number")
    make_hdr(ws_po,["PO Number","Order # (BOL ID)","BOL Number","Carrier",
        "Pallets for this PO","Total Pallets","Weight","Ship To"],[18,12,14,18,16,12,14,28])
    po_map={}
    for bol in bols:
        for j,po in enumerate(bol.get("po_numbers",[])):
            if not po.strip(): continue
            po_map.setdefault(po,[]).append({
                "order_id":bol.get("id",""),"bol_number":bol.get("bol_number",""),
                "carrier":bol.get("carrier",""),
                "pallets_this_po":(bol.get("pallets_per_po") or [])[j] if j<len(bol.get("pallets_per_po",[])) else "",
                "total_pallets":bol.get("total_pallets",""),"weight":bol.get("total_weight",""),
                "address":bol.get("address","")})
    rf=PatternFill("solid",fgColor="FFF3CD"); ri=2
    for po in sorted(po_map):
        entries=po_map[po]; is_rep=len(entries)>1
        for e in entries:
            ws_po.append([po,e["order_id"],e["bol_number"],e["carrier"],
                e["pallets_this_po"],e["total_pallets"],e["weight"],e["address"]])
            style_cells(ws_po,ri,8,rf if is_rep else None,bold=is_rep); ri+=1
    ws_sh=wb.create_sheet("Shortage Sheet")
    make_hdr(ws_sh,["PO Number","Times Ordered","BOL Numbers","Carriers",
        "Total Pallets Across Orders"],[18,14,30,30,22])
    hif=PatternFill("solid",fgColor="F8D7DA"); ri=2
    for po in sorted(po_map):
        entries=po_map[po]; count=len(entries)
        bstr=", ".join(str(e["bol_number"]) for e in entries if e["bol_number"])
        cstr=", ".join(set(str(e["carrier"]) for e in entries if e["carrier"]))
        ptot=sum(int(str(e["pallets_this_po"]).strip()) for e in entries
                 if str(e.get("pallets_this_po","")).strip().isdigit())
        ws_sh.append([po,count,bstr,cstr,ptot if ptot else ""])
        style_cells(ws_sh,ri,5,hif if count>1 else None,bold=(count>1)); ri+=1
    ws_sh.cell(ri+1,1).value="Red rows = PO ordered on multiple BOLs (potential shortage)"
    ws_sh.cell(ri+1,1).font=Font(name="Arial",size=9,italic=True,color="856404")
    tmp=tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False)
    wb.save(tmp.name); tmp.close(); return tmp.name


@app.route("/")
def index(): return render_template("index.html")

@app.route("/api/bols",methods=["GET"])
def list_bols(): return jsonify(load_store())

@app.route("/api/bols",methods=["POST"])
def add_bols():
    data=request.get_json(); bols_input=data.get("bols",[]); store=load_store(); added=[]
    for bd in bols_input:
        c=clean_bol_data(bd); c["id"]=get_next_id(store)
        c["date_added"]=datetime.now().strftime("%Y-%m-%d %H:%M")
        store.append(c); added.append(c)
    save_store(store); return jsonify({"added":len(added),"bols":added})

@app.route("/api/bols/<int:bol_id>",methods=["DELETE"])
def delete_bol(bol_id):
    save_store([b for b in load_store() if b["id"]!=bol_id]); return jsonify({"ok":True})

@app.route("/api/bols/clear",methods=["POST"])
def clear_bols(): save_store([]); return jsonify({"ok":True})

@app.route("/generate",methods=["POST"])
def generate():
    data=request.get_json(); bols=data.get("bols",[])
    if not bols: return jsonify({"error":"No BOL data provided"}),400
    try:
        out=generate_bols_pdf([clean_bol_data(b) for b in bols])
        return send_file(out,mimetype="application/pdf",as_attachment=True,download_name="BOLs.pdf")
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/generate/store",methods=["POST"])
def generate_from_store():
    data=request.get_json(); ids=set(data.get("ids",[])); store=load_store()
    bols=[b for b in store if b["id"] in ids] if ids else store
    if not bols: return jsonify({"error":"No BOLs found"}),400
    try:
        out=generate_bols_pdf(bols)
        return send_file(out,mimetype="application/pdf",as_attachment=True,download_name="BOLs.pdf")
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/import/pdfs",methods=["POST"])
def import_pdfs():
    files=request.files.getlist("files")
    if not files: return jsonify({"error":"No files uploaded"}),400
    results,errors=[],[]
    for f in files:
        if not f.filename.lower().endswith(".pdf"): errors.append(f"{f.filename}: not a PDF"); continue
        tmp=tempfile.NamedTemporaryFile(suffix=".pdf",delete=False)
        try:
            f.save(tmp.name); tmp.close()
            ext=extract_bol_from_pdf(tmp.name); ext["source_filename"]=f.filename; results.append(ext)
        except Exception as e: errors.append(f"{f.filename}: {str(e)}")
        finally:
            try: os.unlink(tmp.name)
            except: pass
    return jsonify({"extracted":results,"errors":errors})

@app.route("/export/excel",methods=["POST"])
def export_excel():
    data=request.get_json(); ids=set(data.get("ids",[])) if data and data.get("ids") else None
    store=load_store(); bols=[b for b in store if b["id"] in ids] if ids else store
    if not bols: return jsonify({"error":"No BOLs to export"}),400
    return send_file(build_excel_shortage(bols),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,download_name="BOL_Shortage_Sheet.xlsx")

if __name__=="__main__":
    app.run(host="0.0.0.0",port=5001,debug=True)
